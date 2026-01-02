"""
Excel handling functions for data exchange.
"""
import io
from typing import List, Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet


# Excel styling constants
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="D4A017", end_color="D4A017", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
CELL_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def auto_size_columns(ws: Worksheet, max_width: int = 50) -> None:
    """Auto-size worksheet columns based on content."""
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except (TypeError, AttributeError):
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, max_width)


def write_excel_header(ws: Worksheet, headers: List[str]) -> None:
    """Write styled header row to Excel worksheet."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = CELL_BORDER


def write_excel_row(ws: Worksheet, row_num: int, values: List[Any]) -> None:
    """Write data row to Excel worksheet with borders."""
    for col, value in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.border = CELL_BORDER


def create_excel_workbook(title: str, headers: List[str], rows: List[List[Any]]) -> bytes:
    """
    Create an Excel workbook with styled header and data rows.

    Args:
        title: Worksheet title
        headers: List of header names
        rows: List of row data

    Returns:
        Excel file as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title

    write_excel_header(ws, headers)
    for row_num, row_data in enumerate(rows, 2):
        write_excel_row(ws, row_num, row_data)

    auto_size_columns(ws)
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def read_excel_rows(content: bytes):
    """
    Read rows from Excel file.

    Args:
        content: Excel file content as bytes

    Yields:
        Tuple of (row_number, row_dict) for each data row
    """
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.rows)]
    header_map = {h: i for i, h in enumerate(headers) if h}

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        row_dict = {}
        for header, idx in header_map.items():
            if idx < len(row):
                row_dict[header] = row[idx]
        yield row_num, row_dict
